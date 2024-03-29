import re
from datetime import date
from bs4 import BeautifulSoup
import requests
from requests.exceptions import ConnectionError, Timeout
from openpyxl import Workbook


class DataNotFoundException(Exception):
    pass


INPUT_FILE = "urls.txt"


class VolleyBallPage:
    """Class to get data off a volleyball webpage

    Raises:
        TypeError: Couldn't find something
        DataNotFoundException: Couldn't find something
        DataNotFoundException: couldn't find something
        TypeError: couldn't find something
        TypeError: couldn't find something

    Returns:
        VolleyBallPage: VolleyBallPage Object
    """

    result_data_re = re.compile(r"\((\w)\).*(\d).*(\d)")
    in_progress_terms = ["In Progress", "Preview Match"]
    DEFAULT_DATA_SIZE = 4

    def __init__(self, url):
        try:
            response = requests.get(url, timeout=20)
        except Timeout:
            response = requests.get(url, timeout=20)
        self.soup = BeautifulSoup(response.content, "html.parser")
        self.url = url

    def _format_default(self, match_result, our_team, other_team, score1, score2):
        """
        Helper method to format a regular data row on the volleyball website

        Raises:
            TypeError: Match Result is not "FFW", "FFL", "W", "L", or "T"
        """
        if match_result == "FFW":
            return (our_team, "N/A", other_team, "N/A")
        elif match_result == "FFL":
            return (other_team, "N/A", our_team, "N/A")
        if match_result == "W":
            our_score = max(score1, score2)
            other_score = min(score1, score2)
            return (our_team, our_score, other_team, other_score)
        elif match_result == "L":
            our_score = min(score1, score2)
            other_score = max(score1, score2)
            return (other_team, other_score, our_team, our_score)
        elif match_result == "T":
            our_score = min(score1, score2)
            other_score = max(score1, score2)
            return (other_team, other_score, our_team, our_score)
        else:
            raise TypeError(f"Game result is not a win or a loss {match_result}")

    def _format_match_TBD(self, our_team, other_team):
        """Formats a match that is pending and has not currently happened

        Args:
            our_team (str): the name of the main team
            other_team (str): the name of the other team

        Returns:
            tuple: formatted in the format: winning team / score / other team / score;
            scores are placeholders
        """
        return (our_team, "Not yet determined", other_team, "Not yet determined")

    def _get_table_rows(self):
        """Gets the row elements from the volleyball web page

        Returns:
            Array[Tag]: list of all rows from a table, including incomplete entries.
        """
        table = self.soup.find("tbody")
        return table.find_all("tr", recursive=False)

    def get_main_team_name(self):
        """Gets the name of the main team for tbe website

        Raises:
            DataNotFoundException: Cannot find the name of the team on the website

        Returns:
            str: Name of the team
        """
        try:
            return re.match(
                "(.*) Volleyball", self.soup.find(id="Team_highlight_info1_Header").text
            ).group(1)
        except AttributeError as a_e:
            raise DataNotFoundException("Cannot find main team's name") from a_e

    def _get_other_team_name(self, trow):
        """Gets the name of the other team from a row

        Args:
            trow (Tag): A row tag from the volleyball website table

        Raises:
            DataNotFoundException: Cannot find other team's name

        Returns:
            str: Other team's name
        """
        try:
            return trow.select_one(".contest-type-indicator").find(
                text=True, recursive=False
            )
        except AttributeError as a_e:
            raise DataNotFoundException("Cannot find other team's name") from a_e

    def _get_score_data(self, trow):
        """Gets the score data from a table row from the volleyball website

        Args:
            trow (Tag): table row from volleyball website

        Raises:
            DataNotFoundException: If that data could not be found
            DataNotFoundException: If the data could not be extracted

        Returns:
            Returns None if a match is pending
            tuple(str, int, int): Tuple containing:
            result (Win, loss, tie) of a team, score of the first team, score of the second team
        """
        result = trow.select_one(".score")
        if result is None:
            for term in VolleyBallPage.in_progress_terms:
                if term in trow.select_one(".last").text:
                    return None
            raise DataNotFoundException(
                f"Data could not be found and a preview match was not determined. {trow.get_text()}"
            )

        data = VolleyBallPage.result_data_re.match(result.text)
        if not data:
            if "FFW" in result.text:
                return ("FFW", "N/A", "N/A")
            elif "FFL" in result.text:
                return ("FFL", "N/A", "N/A")
            raise DataNotFoundException(f"Data could not be extracted {result.text}")

        match_result = data.group(1)
        score1 = int(data.group(2))
        score2 = int(data.group(3))
        return match_result, score1, score2

    def get_volleyball_data(self):
        """Returns a list of data from the volleyball website in the format:
         Winning team / score / Losing team / score

        Returns:
            Array[tuple()]: volleyball data in the format:
             Winning team / score / Losing team / score
        """
        our_team = self.get_main_team_name()
        return_data = []

        for trow in self._get_table_rows():
            try:
                other_team = self._get_other_team_name(trow)
                row_data = self._get_score_data(trow)
                if row_data is not None:
                    match_result, score1, score2 = row_data
                    return_data.append(
                        self._format_default(
                            match_result, our_team, other_team, score1, score2
                        )
                    )
                else:
                    return_data.append(self._format_match_TBD(our_team, other_team))
            except DataNotFoundException as dnfe:
                print(dnfe)
                print(self.url)
                print("skipping...")
                continue
        return return_data


class DataWriter:
    """Represents an excel workbook and has specialized methods for writing volleyball data"""

    def __init__(self) -> None:
        self.book = Workbook()
        self.book.remove(self.book.active)

    def add_volleyball_data(self, sheet_name, data):
        """Adds volleyball data to an excel sheet named sheet_name in the format:
         Winner / score / loser / score

        Args:
            sheet_name (str): Name of the main team
            data (Array(tuple(str, int, str, int))):
            Volleyball data formatted: Winner / score / loser / score
        """
        ws = self.book.create_sheet(sheet_name)
        for i, row in enumerate(
            ws.iter_rows(
                min_row=1, max_row=len(data), max_col=VolleyBallPage.DEFAULT_DATA_SIZE
            )
        ):
            for j, cell in enumerate(row):
                cell.value = data[i][j]

    def save(self):
        """Saves the excel file to disk"""
        book_file_name = f"VolleyballData{date.today().isoformat()}.xlsx"
        self.book.save(book_file_name)


def main():
    """Entrypoint of program"""
    scores = []
    writer = DataWriter()
    with open(INPUT_FILE, "r", encoding="utf8") as reader:
        for url in reader:
            url = url.strip()
            try:
                page = VolleyBallPage(url)
            except ConnectionError as c_e:
                print(f"Website would not connect: {url}")
                print(c_e)
                continue
            except Timeout as t_e:
                print(f"Timed out while trying to connect to website: {url}")
                print(t_e)
                continue
            for element in page.get_volleyball_data():
                if isinstance(element, tuple):
                    scores.append(element)
                else:
                    print(f"invalid value added {element}")
    writer.add_volleyball_data("VolleyballData", scores)
    writer.save()


if __name__ == "__main__":
    main()
